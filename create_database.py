import os
import psycopg2
from openpyxl import load_workbook

# PostgreSQL connection parameters
db_params = {
    'dbname': 'db1',
    'user': 'enid',
    'password': 'enid',
    'host': 'localhost',
}

# Path to your Excel file on Windows desktop
excel_file_path = os.path.join('table-data.xlsx')

create_table_query = """
CREATE TABLE IF NOT EXISTS TableData (
    Date DATE, 
    Chart_Type TEXT, 
    Customer_Id INT, 
    Customer_Name TEXT, 
    Days_Overdue INT, 
    Amount_Outstanding INT, 
    Recovery INT
);
"""
# Connect to PostgreSQL
conn = psycopg2.connect(**db_params)
cursor = conn.cursor()

# Create the table if it doesn't exist
cursor.execute(create_table_query)
conn.commit()

# Load Excel file
wb = load_workbook(excel_file_path)
sheet = wb.active

# Loop through rows and insert data into PostgreSQL
for row in sheet.iter_rows(min_row=2, values_only=True):
    Date, Chart_Type, Customer_Id, Customer_Name, Days_Overdue, Amount_Outstanding, Recovery = row

    # Check for empty cells and perform data transformation if needed
    if Recovery is None:
        Recovery = 0  
    if Days_Overdue is None:
        Days_Overdue = 0  

    insert_query = "INSERT INTO TableData (Date, Chart_Type, Customer_Id, Customer_Name, Days_Overdue, Amount_Outstanding, Recovery) VALUES (%s, %s, %s, %s, %s, %s, %s);"
    cursor.execute(insert_query, (Date, Chart_Type, Customer_Id, Customer_Name, Days_Overdue, Amount_Outstanding, Recovery))

# Commit changes and close connection
conn.commit()
cursor.close()
conn.close()
